import os
import json
import glob
import requests
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from webdriver_manager.chrome import ChromeDriverManager
from datetime import datetime
import gspread
from google.oauth2.service_account import Credentials
import openpyxl
import time

USERNAME      = os.environ["CSI_USERNAME"]
PASSWORD      = os.environ["CSI_PASSWORD"]
GOOGLE_CREDS  = os.environ["GOOGLE_CREDENTIALS"]
SHEET_ID      = "11HKDlLqz4hedo3HWtxNHXHHL8gPS1oN8NlCH_EV5ZfU"
LOGIN_URL     = "https://csi-bdms-mgrs.azurewebsites.net"

def export_excel():
    download_dir = "/tmp/downloads"
    os.makedirs(download_dir, exist_ok=True)

    options = webdriver.ChromeOptions()
    options.add_argument("--headless=new")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    prefs = {
        "download.default_directory": download_dir,
        "download.prompt_for_download": False,
    }
    options.add_experimental_option("prefs", prefs)

    driver = webdriver.Chrome(
        service=Service(ChromeDriverManager().install()), options=options
    )

    try:
        wait = WebDriverWait(driver, 20)

        driver.get(LOGIN_URL)
        wait.until(EC.presence_of_element_located((By.NAME, "username")))
        driver.find_element(By.NAME, "username").send_keys(USERNAME)
        driver.find_element(By.NAME, "password").send_keys(PASSWORD)
        driver.find_element(By.ID, "login").click()
        wait.until(EC.url_contains("FirstPage"))
        print("Login success")

        driver.get(f"{LOGIN_URL}/Home/Export?uid=87")
        time.sleep(3)
        print("Export page loaded")

        wait.until(EC.presence_of_element_located((By.TAG_NAME, "select")))
        select = Select(driver.find_element(By.TAG_NAME, "select"))
        select.select_by_visible_text("BHP")
        print("Selected BHP")
        time.sleep(2)

        labels = driver.find_elements(By.CSS_SELECTOR, "label")
        for label in labels:
            text = label.text.strip()
            if "(" in text and ")" in text:
                try:
                    checkbox_id = label.get_attribute("for")
                    checkbox = driver.find_element(By.ID, checkbox_id)
                    if not checkbox.is_selected():
                        checkbox.click()
                    print("Ticked: " + text)
                except:
                    pass
        time.sleep(1)

        export_btn = wait.until(EC.element_to_be_clickable(
            (By.CLASS_NAME, "btn-success")
        ))
        export_btn.click()
        print("Clicked Export")
        time.sleep(5)

        files = glob.glob(f"{download_dir}/*.xlsx")
        if not files:
            files = glob.glob(f"{download_dir}/*")
        if files:
            filepath = max(files, key=os.path.getctime)
            print("Downloaded: " + filepath)
            return filepath
        else:
            print("No file found")
            return None

    finally:
        driver.quit()

def upload_to_sheets(filepath):
    creds_dict = json.loads(GOOGLE_CREDS)
    creds = Credentials.from_service_account_info(
        creds_dict,
        scopes=["https://www.googleapis.com/auth/spreadsheets"]
    )
    gc = gspread.authorize(creds)
    sh = gc.open_by_key(SHEET_ID)

    wb = openpyxl.load_workbook(filepath)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        data = []
        for row in ws.iter_rows(values_only=True):
            data.append([str(cell) if cell is not None else "" for cell in row])

        try:
            worksheet = sh.worksheet(sheet_name)
            worksheet.clear()
        except:
            worksheet = sh.add_worksheet(title=sheet_name, rows=1000, cols=20)

        worksheet.update(data)
        print("Updated sheet: " + sheet_name)

    return "https://docs.google.com/spreadsheets/d/" + SHEET_ID

filepath = export_excel()
if filepath:
    upload_to_sheets(filepath)
else:
    print("No file found")
