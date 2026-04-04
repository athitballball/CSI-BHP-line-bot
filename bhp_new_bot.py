import os
import json
import glob
import time
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from webdriver_manager.chrome import ChromeDriverManager
import gspread
from google.oauth2.service_account import Credentials
import openpyxl

# แก้ไขเครื่องหมายคำพูดจาก “ ” เป็น " "
USERNAME     = os.environ.get("CSI_USERNAME")
PASSWORD     = os.environ.get("CSI_PASSWORD")
GOOGLE_CREDS = os.environ.get("GOOGLE_CREDENTIALS")
SHEET_ID     = "11dX9ga5X5yZBeL-Nb__F1bIS96QdIVbPZJ93QX7e0_E"
LOGIN_URL    = "https://csi-bdms-mgrs.azurewebsites.net"
START_DATE   = "01/Mar/2026"

def export_excel():
    download_dir = "/tmp/downloads"
    os.makedirs(download_dir, exist_ok=True)

    options = webdriver.ChromeOptions()
    options.add_argument("--headless=new")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--window-size=1920,1080")
    options.add_experimental_option("prefs", {
        "download.default_directory": download_dir,
        "download.prompt_for_download": False,
    })

    driver = webdriver.Chrome(
        service=Service(ChromeDriverManager().install()), options=options
    )

    try:
        wait = WebDriverWait(driver, 20)

        driver.get(LOGIN_URL)
        wait.until(EC.presence_of_element_located((By.NAME, "username")))
        driver.find_element(By.NAME, "username").send_keys(USERNAME)
        driver.find_element(By.NAME, "password").send_keys(PASSWORD)
        driver.find_element(By.ID, "login").click()
        wait.until(EC.url_contains("FirstPage"))
        print("Login success")

        driver.get(LOGIN_URL + "/Home/Export?uid=87")
        time.sleep(3)
        print("Export page loaded")

        wait.until(EC.presence_of_element_located((By.TAG_NAME, "select")))
        Select(driver.find_element(By.TAG_NAME, "select")).select_by_visible_text("BHP")
        print("Selected BHP")
        time.sleep(2)

        # ค้นหา input วันที่
        date_inputs = driver.find_elements(By.XPATH, "//label[contains(text(),'\u0e27\u0e31\u0e19\u0e17\u0e35\u0e48\u0e40\u0e23\u0e34\u0e48\u0e21\u0e15\u0e49\u0e19')]/..//input | //label[contains(text(),'\u0e27\u0e31\u0e19\u0e17\u0e35\u0e48\u0e2a\u0e34\u0e49\u0e19\u0e2a\u0e38\u0e14')]/..//input")
        if len(date_inputs) < 2:
            all_inputs = driver.find_elements(By.CSS_SELECTOR, "input[type='text']")
            date_inputs = all_inputs[1:3]

        # ใส่ Start Date
        ActionChains(driver).move_to_element(date_inputs[0]).click().perform()
        time.sleep(1)
        date_inputs[0].send_keys(Keys.CONTROL + "a")
        date_inputs[0].send_keys(START_DATE)
        date_inputs[0].send_keys(Keys.ENTER)
        time.sleep(2)
        print(f"Start date typed: {START_DATE}")

        # ใส่ End Date
        end_date = datetime.now().strftime("%d/%b/%Y")
        ActionChains(driver).move_to_element(date_inputs[1]).click().perform()
        time.sleep(1)
        date_inputs[1].send_keys(Keys.CONTROL + "a")
        date_inputs[1].send_keys(end_date)
        date_inputs[1].send_keys(Keys.ENTER)
        time.sleep(10)
        print(f"End date typed: {end_date}")

        # ติ๊ก Checkbox ต่างๆ
        labels = driver.find_elements(By.CSS_SELECTOR, "label")
        for label in labels:
            text = label.text.strip()
            if "(" in text and ")" in text:
                try:
                    checkbox_id = label.get_attribute("for")
                    checkbox = driver.find_element(By.ID, checkbox_id) if checkbox_id else label.find_element(By.TAG_NAME, "input")
                    if not checkbox.is_selected():
                        driver.execute_script("arguments[0].click();", checkbox)
                    print(f"Ticked: {text}")
                except Exception as e:
                    print(f"Could not tick: {text} -> {e}")
        
        time.sleep(1)
        driver.execute_script("document.body.click();")
        time.sleep(1)

        export_btn = wait.until(EC.presence_of_element_located((By.ID, "exportBtn")))
        driver.execute_script("arguments[0].click();", export_btn)
        print("Clicked Export")
        time.sleep(10)

        # ค้นหาไฟล์ที่โหลดมา
        files = glob.glob(download_dir + "/*.xlsx") or glob.glob(download_dir + "/*")
        if files:
            filepath = max(files, key=os.path.getctime)
            print(f"Downloaded: {filepath}")
            return filepath
        else:
            print("No file found")
            return None

    finally:
        driver.quit()

def upload_to_sheets(filepath):
    if not GOOGLE_CREDS:
        print("Google Credentials not found in environment")
        return

    creds = Credentials.from_service_account_info(
        json.loads(GOOGLE_CREDS),
        scopes=["https://www.googleapis.com/auth/spreadsheets"]
    )
    gc = gspread.authorize(creds)
    sh = gc.open_by_key(SHEET_ID)
    wb = openpyxl.load_workbook(filepath)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        data = [
            [str(cell) if cell is not None else "" for cell in row]
            for row in ws.iter_rows(values_only=True)
        ]
        try:
            worksheet = sh.worksheet(sheet_name)
            worksheet.clear()
        except Exception:
            worksheet = sh.add_worksheet(title=sheet_name, rows=5000, cols=30)

        worksheet.update('A1', data) # ระบุ Range เริ่มต้น
        print(f"Updated sheet: {sheet_name}")

    print(f"Success: https://docs.google.com/spreadsheets/d/{SHEET_ID}")

# Main execution
if __name__ == "__main__":
    filepath = export_excel()
    if filepath:
        upload_to_sheets(filepath)
    else:
        print("Export failed")
